from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import openpyxl
import time
import os
from time import sleep
import glob
import pyautogui
import requests

API_KEY = "62a06978600e98d3cbf430ffe18ea254"

def registrar_empresa_sem_debitos(nome_empresa, caminho_arquivo="empresas_sem_debitos.xlsx"):
    # Verifica se o arquivo já existe
    if not os.path.exists(caminho_arquivo):
        # Cria um novo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas Sem Débitos"
        # Adiciona cabeçalhos
        ws.append(["Nome da Empresa", "Mensagem"])
    else:
        # Carrega o arquivo existente
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active

    # Adiciona os dados da empresa
    ws.append([nome_empresa, "empresa sem débitos"])

    # Salva o arquivo
    wb.save(caminho_arquivo)
    print(f"Empresa '{nome_empresa}' registrada no Excel.")

def capturar_tela_inteira(caminho_salvar):
    # Captura a tela inteira
    screenshot = pyautogui.screenshot()
    # Salva a imagem no caminho especificado
    screenshot.save(caminho_salvar)


def capturar_regiao_captcha(x1, y1, x2, y2, output_path):
    try:
        largura = x2 - x1  # Calcula a largura do retângulo
        altura = y2 - y1   # Calcula a altura do retângulo
        print(f"Capturando região com largura={largura} e altura={altura}")
        
        # Captura apenas a região especificada
        screenshot = pyautogui.screenshot(region=(x1, y1, largura, altura))
        screenshot.save(output_path)
        print(f"CAPTCHA salvo em: {output_path}")
        return output_path
    except Exception as e:
        print(f"Erro ao capturar a região do CAPTCHA: {e}")
        return None
    
def resolver_captcha_2captcha(image_path):
    try:
        # Passo 1: Enviar a imagem do CAPTCHA
        print("Enviando CAPTCHA para o 2Captcha...")
        with open(image_path, 'rb') as img_file:
            files = {'file': img_file}
            response = requests.post(
                f"http://2captcha.com/in.php?key={API_KEY}&method=post",
                files=files
            )
        
        if response.status_code != 200 or "OK" not in response.text:
            raise Exception("Erro ao enviar o CAPTCHA para o 2Captcha")
        
        captcha_id = response.text.split('|')[1]
        print(f"CAPTCHA enviado. ID: {captcha_id}")
        
        # Passo 2: Aguardar e consultar o resultado
        for _ in range(30):  # Tenta por até 30 segundos
            time.sleep(5)  # Espera 5 segundos entre consultas
            result_response = requests.get(f"http://2captcha.com/res.php?key={API_KEY}&action=get&id={captcha_id}")
            if result_response.text == "CAPCHA_NOT_READY":
                print("CAPTCHA ainda não está pronto, aguardando...")
                continue
            
            if "OK" in result_response.text:
                captcha_text = result_response.text.split('|')[1]
                print(f"CAPTCHA resolvido: {captcha_text}")
                return captcha_text
            
        raise Exception("Tempo esgotado para resolver o CAPTCHA")
    except Exception as e:
        print(f"Erro ao resolver CAPTCHA com 2Captcha: {e}")
        return None
    
def pegar_débitos_sp():
    # Configura o caminho da pasta de download automático
    download_path = os.path.abspath("Boletos")
    os.makedirs(download_path, exist_ok=True)

    # Configuração das opções do Chrome
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--ignore-certificate-errors")

    # Configura o download automático
    prefs = {
        "download.default_directory": download_path,  # Define a pasta de downloads
        "download.prompt_for_download": False,  # Não perguntar onde salvar
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,  # Proteção contra downloads perigosos
        "profile.default_content_settings.popups": 0
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # Inicializa o driver
    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.get('https://senhawebsts.prefeitura.sp.gov.br/Account/Login.aspx?ReturnUrl=%2f%3fwa%3dwsignin1.0%26wtrealm%3dhttps%253a%252f%252fduc.prefeitura.sp.gov.br%252fportal%252f%26wctx%3drm%253d0%2526id%253dpassive%2526ru%253d%25252fportal%25252f%26wct%3d2024-12-16T11%253a48%253a11Z&wa=wsignin1.0&wtrealm=https%3a%2f%2fduc.prefeitura.sp.gov.br%2fportal%2f&wctx=rm%3d0%26id%3dpassive%26ru%3d%252fportal%252f&wct=2024-12-16T11%3a48%253a11Z')


    wb = openpyxl.load_workbook('dados_sp.xlsx')
    sheet_wb = wb['São Paulo']
    try:
        for indice, linha in enumerate(sheet_wb.iter_rows(min_row=9, max_row=10)):  # Ajuste o intervalo conforme necessário
            login = linha[4].value
            senha = linha[5].value
            ccm = linha[3].value
            nome_empresa = linha[0].value

            # Preencher Login
            clicar_login = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$txtUser']"))
            )
            clicar_login.send_keys(str(login))

            sleep(2)

            # Preencher Senha
            clicar_senha = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$txtPassword']"))
            )
            clicar_senha.send_keys(str(senha))
            
            try:
                # Captura a região do CAPTCHA
                captcha_path = "captcha_regiao.png"
                capturar_regiao_captcha(428, 455, 780, 600, captcha_path)  # regiao exata
                
                # Resolver o CAPTCHA via 2Captcha
                captcha_resposta = resolver_captcha_2captcha(captcha_path)
                print(captcha_resposta)
                if not captcha_resposta:
                    print("Falha ao resolver o CAPTCHA.")
                    return
                
                # Preencher o formulário
                campo_captcha = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,"//input[@name='ctl00$ctl00$formBody$formBody$wucRecaptcha1$txtValidacao']"))
                )
                campo_captcha.click()
                campo_captcha.send_keys(captcha_resposta)  
                
                print("CAPTCHA preenchido com sucesso.")
                
                botao_submit = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$btnLogin']"))
                )
                botao_submit.click()

            except Exception as e:
                print('Captcha não encontrado')
            
            
            sleep(3)
            botao_mais = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//img[@id='img_maisTwo']"))
            )
            botao_mais.click()

            sleep(1)

            botao_ccm = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@name='ctl00$ctl00$ConteudoPrincipal$ContentPlaceHolder1$txtCCM']"))
            )
            botao_ccm.click()
            botao_ccm.send_keys(ccm)

            sleep(1)

            botao_OK = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@name='ctl00$ctl00$ConteudoPrincipal$ContentPlaceHolder1$btnOKDAMSP']"))
            )

            botao_OK.click()

            sleep(3)

            
            try:
                # Caminho onde a imagem do CAPTCHA será salva
                caminho_captcha = "captcha_tela_inteira.png"
                
                
                capturar_regiao_captcha(0, 0, 515, 250, caminho_captcha)
                
                # Resolver o CAPTCHA via 2Captcha
                captcha_resposta_2 = resolver_captcha_2captcha(caminho_captcha)
                print(captcha_resposta_2)
                
                if not captcha_resposta_2:
                    print("Falha ao resolver o CAPTCHA.")
                    return
            except Exception as e:
                print(f"Ocorreu um erro: {e}")

            campo = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@id='ans']"))
            )
            campo.click()
            campo.send_keys(captcha_resposta_2)
            

            submit = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@id='jar']"))
            )
            submit.click()
            

            #print(driver.page_source)

            driver.switch_to.default_content()
            sleep(5)
            try:
                pagar = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//input[@name='ctl00$ctl00$ConteudoPrincipal$ContentPlaceHolder1$btnPagar']"))
                )

                pagar.click()
                try:
                    alert = Alert(driver)
                    alert.accept()  # Fecha o pop-up clicando em "OK"

                except Exception:
                    print("Nenhum alerta foi encontrado.")
                
                sleep(2)
                pyautogui.hotkey('ctrl', 'w')

                WebDriverWait(driver, 10).until(lambda driver: len(driver.window_handles) > 0)

                # Alternar para a última janela que permanece aberta
                driver.switch_to.window(driver.window_handles[-1])
                driver.get('https://senhawebsts.prefeitura.sp.gov.br/Account/Login.aspx?ReturnUrl=%2f%3fwa%3dwsignin1.0%26wtrealm%3dhttps%253a%252f%252fduc.prefeitura.sp.gov.br%252fportal%252f%26wctx%3drm%253d0%2526id%253dpassive%2526ru%253d%25252fportal%25252f%26wct%3d2024-12-16T11%253a48%253a11Z&wa=wsignin1.0&wtrealm=https%3a%2f%2fduc.prefeitura.sp.gov.br%2fportal%2f&wctx=rm%3d0%26id%3dpassive%26ru%3d%252fportal%252f&wct=2024-12-16T11%3a48%253a11Z')
            
            except Exception as e:
                nome_empresa = linha[0].value  # Captura o nome da empresa na planilha original
                print(f"Empresa '{nome_empresa}' sem débitos.")
                registrar_empresa_sem_debitos(nome_empresa)
                driver.get('https://senhawebsts.prefeitura.sp.gov.br/Account/Login.aspx?ReturnUrl=%2f%3fwa%3dwsignin1.0%26wtrealm%3dhttps%253a%252f%252fduc.prefeitura.sp.gov.br%252fportal%252f%26wctx%3drm%253d0%2526id%253dpassive%2526ru%253d%25252fportal%25252f%26wct%3d2024-12-16T11%253a48%253a11Z&wa=wsignin1.0&wtrealm=https%3a%2f%2fduc.prefeitura.sp.gov.br%2fportal%2f&wctx=rm%3d0%26id%3dpassive%26ru%3d%252fportal%252f&wct=2024-12-16T11%3a48%253a11Z')

            arquivo_antigo = aguardar_download(download_path)
            if arquivo_antigo:
                novo_nome = f"{nome_empresa}.pdf"  # Nome da empresa + extensão
                novo_caminho = os.path.join(download_path, novo_nome)
                os.rename(arquivo_antigo, novo_caminho)
                print(f"Arquivo salvo como: {novo_nome}")
            else:
                print("Nenhum arquivo encontrado para renomear.")

        
            time.sleep(5)
        
    finally:
        driver.quit()

def aguardar_download(diretorio, timeout=60):
    """ Aguarda o download do arquivo terminar e retorna o caminho do arquivo baixado. """
    tempo_inicial = time.time()
    while time.time() - tempo_inicial < timeout:
        arquivos = glob.glob(os.path.join(diretorio, "*.pdf"))  # Busca arquivos PDF
        arquivos_em_andamento = glob.glob(os.path.join(diretorio, "*.crdownload"))  # Arquivos incompletos
        if arquivos and not arquivos_em_andamento:
            return arquivos[0]  # Retorna o primeiro arquivo PDF encontrado
        time.sleep(1)
    return None

# Executa a função
pegar_débitos_sp()
