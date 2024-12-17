from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import openpyxl
import time
import os
from time import sleep
import glob
import pyautogui
import requests
from selenium.common.exceptions import TimeoutException
import threading
import customtkinter as ctk
from PIL import Image, ImageTk


API_KEY = "62a06978600e98d3cbf430ffe18ea254"
caminho_imagem = 'image.png'
ultima_linha_processada = ""

def ler_progresso_sao_paulo():
    if os.path.exists("progresso_sao_paulo.txt"):
        with open("progresso_sao_paulo.txt", "r") as file:
            return int(file.read().strip())
    return 2  # Se não houver progresso registrado, começa da linha 2

# Função para salvar o progresso de São Paulo
def salvar_progresso_sao_paulo(linha):
    with open("progresso_sao_paulo.txt", "w") as file:
        file.write(str(linha))

# Função para atualizar a última linha processada de São Paulo
def atualizar_ultima_linha_sao_paulo():
    ultima_linha_sao_paulo = ler_progresso_sao_paulo()
    ultima_sao_paulo.configure(text=f"(Última linha processada: {ultima_linha_sao_paulo})")

def atualizar_ultima_linha():
    # Atualiza o texto do label com a última linha processada
    ultima_sao_paulo.configure(text=f"(Última linha processada: {ultima_linha_processada_sao_paulo})")
    # Agenda a atualização para o próximo intervalo (em milissegundos)
    app.after(1000, atualizar_ultima_linha)  # Atualiza a cada 1 segundo (1000 ms)
    

def registrar_empresa_sem_debitos(nome_empresa, caminho_arquivo="empresas_sem_debitos.xlsx"):
    # Verifica se o arquivo já existe
    if not os.path.exists(caminho_arquivo):
        # Cria um novo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas Sem Débitos"
        # Adiciona cabeçalhos
        ws.append(["Nome da Empresa", "Mensagem"])
    else:
        # Carrega o arquivo existente
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active

    # Adiciona os dados da empresa
    ws.append([nome_empresa, "empresa sem débitos"])

    # Salva o arquivo
    wb.save(caminho_arquivo)
    print(f"Empresa '{nome_empresa}' registrada no Excel.")

def capturar_tela_inteira(caminho_salvar):
    # Captura a tela inteira
    screenshot = pyautogui.screenshot()
    # Salva a imagem no caminho especificado
    screenshot.save(caminho_salvar)


def capturar_regiao_captcha(x1, y1, x2, y2, output_path):
    try:
        largura = x2 - x1  # Calcula a largura do retângulo
        altura = y2 - y1   # Calcula a altura do retângulo
        print(f"Capturando região com largura={largura} e altura={altura}")
        
        # Captura apenas a região especificada
        screenshot = pyautogui.screenshot(region=(x1, y1, largura, altura))
        screenshot.save(output_path)
        print(f"CAPTCHA salvo em: {output_path}")
        return output_path
    except Exception as e:
        print(f"Erro ao capturar a região do CAPTCHA: {e}")
        return None
    
def resolver_captcha_2captcha(image_path):
    try:
        # Passo 1: Enviar a imagem do CAPTCHA
        print("Enviando CAPTCHA para o 2Captcha...")
        with open(image_path, 'rb') as img_file:
            files = {'file': img_file}
            response = requests.post(
                f"http://2captcha.com/in.php?key={API_KEY}&method=post",
                files=files
            )
        
        if response.status_code != 200 or "OK" not in response.text:
            raise Exception("Erro ao enviar o CAPTCHA para o 2Captcha")
        
        captcha_id = response.text.split('|')[1]
        print(f"CAPTCHA enviado. ID: {captcha_id}")
        
        # Passo 2: Aguardar e consultar o resultado
        for _ in range(45):  # Tenta por até 30 segundos
            time.sleep(5)  # Espera 5 segundos entre consultas
            result_response = requests.get(f"http://2captcha.com/res.php?key={API_KEY}&action=get&id={captcha_id}")
            if result_response.text == "CAPCHA_NOT_READY":
                print("CAPTCHA ainda não está pronto, aguardando...")
                continue
            
            if "OK" in result_response.text:
                captcha_text = result_response.text.split('|')[1]
                print(f"CAPTCHA resolvido: {captcha_text}")
                return captcha_text
            
        raise Exception("Tempo esgotado para resolver o CAPTCHA")
    except Exception as e:
        print(f"Erro ao resolver CAPTCHA com 2Captcha: {e}")
        return None
    
def verificar_captcha_2_completo(driver,caminho_captcha):
    try:
        # Caminho onde a imagem do CAPTCHA será salva
        caminho_captcha = "captcha_tela_inteira.png"

        capturar_regiao_captcha(0, 0, 515, 250, caminho_captcha)
        
        # Resolver o CAPTCHA via 2Captcha
        captcha_resposta_2 = resolver_captcha_2captcha(caminho_captcha)
        print(f"Resposta do CAPTCHA: {captcha_resposta_2}")
        
        if not captcha_resposta_2:
            print("Falha ao resolver o CAPTCHA.")
            return
    except Exception as e:
        print(f"Ocorreu um erro ao resolver o CAPTCHA: {e}")

    # Enviar a resposta para o campo
    campo = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@id='ans']"))
    )
    campo.clear()  # Limpa qualquer valor anterior
    campo.send_keys(captcha_resposta_2)

    # Submeter o CAPTCHA
    submit = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@id='jar']"))
    )
    submit.click()
    
    sleep(3)
    # Verificar se o CAPTCHA foi aceito ou deu erro
    try:
        print('Verificando possível erro no captcha 2')
        # Espera até 5 segundos pela mensagem de erro
        mensagem_erro = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body[contains(text(), 'Você digitou o código errado.')]"))
        ) #ta certo!!
        print("O CAPTCHA foi digitado incorretamente. Tentando novamente...")

        # Se a mensagem de erro for encontrada, repetir o processo
        capturar_regiao_captcha(0, 0, 515, 250, caminho_captcha)
        captcha_resposta_2 = resolver_captcha_2captcha(caminho_captcha)
        print(f"Nova resposta do CAPTCHA: {captcha_resposta_2}")
        
        campo = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@id='ans']"))
        )
        campo.clear()
        campo.send_keys(captcha_resposta_2)
        print('Captcha 2 preenchido')
        
        try:
            # Reencontrar o botão de envio antes de clicar
            submit = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@id='jar']"))
            )
            submit.click()

            try:
                print('Verificando possível erro pela 3 vez no captcha 2')
                # Espera até 5 segundos pela mensagem de erro
                mensagem_erro = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//body[contains(text(), 'Você digitou o código errado.')]"))
                ) #ta certo!!
                print("O CAPTCHA foi digitado incorretamente. Tentando novamente...")

                # Se a mensagem de erro for encontrada, repetir o processo
                capturar_regiao_captcha(0, 0, 515, 250, caminho_captcha)
                captcha_resposta_2 = resolver_captcha_2captcha(caminho_captcha)
                print(f"Nova resposta do CAPTCHA: {captcha_resposta_2}")
                
                campo = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@id='ans']"))
                )
                campo.clear()
                campo.send_keys(captcha_resposta_2)
                print('Captcha 2 preenchido pela 3 vez')

                submit = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@id='jar']"))
                )
                submit.click()
            except Exception:
                print('Erro nao encontrado pela 3 vez')

        except Exception:
            print("Elemento do botão de envio se tornou obsoleto. Tentando localizá-lo novamente...")
        
    except TimeoutException:
        print("CAPTCHA resolvido com sucesso.")

def pegar_débitos_sp():
    # Configura o caminho da pasta de download automático
    download_path = os.path.abspath("Boletos")
    os.makedirs(download_path, exist_ok=True)

    # Configuração das opções do Chrome
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--ignore-certificate-errors")

    # Configura o download automático
    prefs = {
        "download.default_directory": download_path,  # Define a pasta de downloads
        "download.prompt_for_download": False,  # Não perguntar onde salvar
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,  # Proteção contra downloads perigosos
        "profile.default_content_settings.popups": 0
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # Inicializa o driver
    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()
    
    wb = openpyxl.load_workbook('dados_sp.xlsx')
    sheet_wb = wb['São Paulo']

    try:
        ultima_linha_processada_sao_paulo = ler_progresso_sao_paulo()
        print(f"Iniciando da linha {ultima_linha_processada_sao_paulo}")
        for indice, linha in enumerate(sheet_wb.iter_rows(min_row=ultima_linha_processada_sao_paulo, max_row=300)): 
            driver.get('https://senhawebsts.prefeitura.sp.gov.br/Account/Login.aspx?ReturnUrl=%2f%3fwa%3dwsignin1.0%26wtrealm%3dhttps%253a%252f%252fduc.prefeitura.sp.gov.br%252fportal%252f%26wctx%3drm%253d0%2526id%253dpassive%2526ru%253d%25252fportal%25252f%26wct%3d2024-12-16T11%253a48%253a11Z&wa=wsignin1.0&wtrealm=https%3a%2f%2fduc.prefeitura.sp.gov.br%2fportal%2f&wctx=rm%3d0%26id%3dpassive%26ru%3d%252fportal%252f&wct=2024-12-16T11%3a48%253a11Z')
            salvar_progresso_sao_paulo(linha[0].row)
            sleep(2)
            login = linha[4].value
            senha = linha[5].value
            ccm = linha[3].value
            nome_empresa = linha[0].value

            if not login or str(login).strip() == "-":
                print(f"Linha {indice + 1}: Login inválido ou ausente. Pulando para a próxima linha.")
                continue  # Pula para a próxima iteração

            # Preencher Login
            clicar_login = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$txtUser']"))
            )
            clicar_login.send_keys(str(login))

            sleep(2)

            # Preencher Senha
            clicar_senha = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$txtPassword']"))
            )
            clicar_senha.send_keys(str(senha))
            
            while True:  # Loop para resolver o CAPTCHA até passar
                try:
                    # Captura a região do CAPTCHA
                    captcha_path = "captcha_regiao.png"
                    capturar_regiao_captcha(428, 455, 780, 600, captcha_path)  # Região exata do CAPTCHA
                    
                    # Resolver o CAPTCHA via 2Captcha
                    captcha_resposta = resolver_captcha_2captcha(captcha_path)
                    
                    if not captcha_resposta:
                        print("Falha ao resolver o CAPTCHA.")
                        continue  # Volta para o início do loop para tentar novamente

                    # Preencher o campo do CAPTCHA
                    campo_captcha = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$wucRecaptcha1$txtValidacao']"))
                    )
                    campo_captcha.clear()  # Limpa o campo caso tenha algo
                    campo_captcha.send_keys(captcha_resposta)

                    # Clicar no botão de login
                    botao_submit = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$btnLogin']"))
                    )
                    botao_submit.click()

                    sleep(3)  # Espera carregar a página após o envio do CAPTCHA

                    # Verificar se a mensagem de erro do CAPTCHA apareceu
                    mensagem_erro = driver.find_elements(By.XPATH, "//span[@id='formBody_formBody_RecaptchaValidator' and contains(@class, 'label-danger')]")
                    if mensagem_erro:
                        print("CAPTCHA incorreto, tentando novamente...")
                        clicar_login = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$txtUser']"))
                        )
                        clicar_login.clear()
                        clicar_login.send_keys(str(login))

                        sleep(2)

                        # Preencher Senha
                        clicar_senha = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, "//input[@name='ctl00$ctl00$formBody$formBody$txtPassword']"))
                        )
                        clicar_senha.clear()
                        clicar_senha.send_keys(str(senha))
                        continue  # Se a mensagem de erro apareceu, tenta resolver novamente
                    else:
                        print("CAPTCHA resolvido com sucesso!")
                        break  # Sai do loop se não houver mensagem de erro

                except Exception as e:
                    print(f"Erro ao tentar resolver o CAPTCHA: {e}")
                    continue

            sleep(5)

            try:
                elemento_presente = WebDriverWait(driver, 4).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@id='ans']"))
                )
                
                # Caso o elemento seja clicável, executa a função
                caminho_captcha = "captcha_tela_inteira.png"  # Defina o caminho do captcha conforme necessário
                verificar_captcha_2_completo(driver, caminho_captcha)
            except Exception as e:
                print("Elemento não encontrado ou não está clicável. Erro:", str(e))
            
            botao_mais = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//img[@id='img_maisTwo']"))
            )
            botao_mais.click()

            sleep(1)

            botao_ccm = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@name='ctl00$ctl00$ConteudoPrincipal$ContentPlaceHolder1$txtCCM']"))
            )
            botao_ccm.click()
            botao_ccm.send_keys(ccm)

            sleep(1)

            botao_OK = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@name='ctl00$ctl00$ConteudoPrincipal$ContentPlaceHolder1$btnOKDAMSP']"))
            )

            botao_OK.click()
            
            try:
                elemento_presente = WebDriverWait(driver, 4).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@id='ans']"))
                )
                
                # Caso o elemento seja clicável, executa a função
                caminho_captcha = "captcha_tela_inteira.png"  # Defina o caminho do captcha conforme necessário
                verificar_captcha_2_completo(driver, caminho_captcha)
            except Exception as e:
                print("Elemento não encontrado ou não está clicável. Erro:", str(e))
        

            sleep(3)
            
            #print(driver.page_source)

            driver.switch_to.default_content()

            try:
                pagar = WebDriverWait(driver,4).until(
                    EC.element_to_be_clickable((By.XPATH,"//input[@name='ctl00$ctl00$ConteudoPrincipal$ContentPlaceHolder1$btnPagar']"))
                )

                pagar.click()
                try:
                    alert = Alert(driver)
                    alert.accept()  # Fecha o pop-up clicando em "OK"

                except Exception:
                    print("Nenhum alerta foi encontrado.")
                
                sleep(2)
                pyautogui.hotkey('ctrl', 'w')

                WebDriverWait(driver, 5).until(lambda driver: len(driver.window_handles) > 0)

                # Alternar para a última janela que permanece aberta
                driver.switch_to.window(driver.window_handles[-1])
                
            
            except Exception as e:
                nome_empresa = linha[0].value  # Captura o nome da empresa na planilha original
                print(f"Empresa '{nome_empresa}' sem débitos.")
                registrar_empresa_sem_debitos(nome_empresa)
                

            arquivo_antigo = aguardar_download(download_path)
            if arquivo_antigo:
                novo_nome = f"{nome_empresa}.pdf"  # Nome da empresa + extensão
                novo_caminho = os.path.join(download_path, novo_nome)
                os.rename(arquivo_antigo, novo_caminho)
                print(f"Arquivo salvo como: {novo_nome}")
            else:
                print("Nenhum arquivo encontrado para renomear.")

            salvar_progresso_sao_paulo(indice + 1)
            time.sleep(2)
            
        
    finally:
        driver.quit()

def aguardar_download(diretorio, timeout=5):
    """ Aguarda o download do arquivo terminar e retorna o caminho do arquivo baixado. """
    tempo_inicial = time.time()
    while time.time() - tempo_inicial < timeout:
        arquivos = glob.glob(os.path.join(diretorio, "*.pdf"))  # Busca arquivos PDF
        arquivos_em_andamento = glob.glob(os.path.join(diretorio, "*.crdownload"))  # Arquivos incompletos
        if arquivos and not arquivos_em_andamento:
            return arquivos[0]  # Retorna o primeiro arquivo PDF encontrado
        time.sleep(1)
    return None

def processar_debitos():
    try:
    # Exemplo de chamada para o seu código que processa as informações de débitos
        pegar_débitos_sp()
        resultado_label.configure(text="Processo concluído com sucesso!", text_color="green")
    except Exception as e:
        resultado_label.configure(text=f"Erro: {str(e)}", text_color="red")

# Função que inicia o processamento em uma thread separada para não travar a interface
def iniciar_processo():
    # Atualiza a interface para indicar que o processo está em andamento
    global ultima_linha_processada

    for i in range(5):
        # Simulação de um processo que faz algo, como buscar dados de uma linha
        ultima_linha_processada = f"Linha {i+1} processada"

        if janela_aberta(app):
            resultado_label.configure(text=f"Processando... {i+1}", text_color="yellow")
            app.update_idletasks()  # Atualiza a interface
            time.sleep(1)  # Simula o tempo de processamento
        else:
            break  # Caso a janela tenha sido fechada, pare o processamento
    if janela_aberta(app):
        resultado_label.configure(text="Processo concluído!", text_color="green")

    if not janela_aberta(app):
        print(f"Última linha processada: {ultima_linha_processada}")

    resultado_label.configure(text="Processando... Aguarde.", text_color="yellow")
    # Obtém o valor do campo de entrada, que é a linha inicial
    try:
        linha_inicial = int(entry_sao_paulo.get())  # Lê a linha inicial a partir da entrada
        if linha_inicial < 2:
            raise ValueError("A linha inicial deve ser maior ou igual a 2")
    except ValueError as e:
        resultado_label.configure(text=f"Erro: {e}", text_color="red")
        return
    
    # Atualiza o progresso e inicia a execução
    salvar_progresso_sao_paulo(linha_inicial)
    atualizar_ultima_linha_sao_paulo()

    # Começa o processo em uma thread separada para não travar a interface
    process_thread = threading.Thread(target=processar_debitos)
    process_thread.start()

def thread_processo():
    # Executa a função iniciar_processo em uma thread separada
    threading.Thread(target=iniciar_processo, daemon=True).start()

def janela_aberta(window):
    return window.winfo_exists()

def fechar_janela():
    global ultima_linha_processada_sao_paulo
    # Atualize a última linha processada antes de fechar a janela
    if ultima_linha_processada_sao_paulo:
        print(f"A janela foi fechada. Última linha processada: {ultima_linha_processada_sao_paulo}")
    app.quit()  # Fecha a janela
def exibir_imagem():
    # Abra a imagem usando Pillow
    imagem = Image.open(caminho_imagem)
    
    # Redimensione a imagem se necessário (opcional)
    imagem = imagem.resize((200, 200))  # Defina o tamanho desejado
    
    # Converta a imagem para um formato que o customtkinter entenda
    imagem_tk = ImageTk.PhotoImage(imagem)
    
    # Exiba a imagem em um widget CTkLabel
    imagem_label = ctk.CTkLabel(app, image=imagem_tk)
    imagem_label.image = imagem_tk  # Mantenha uma referência à imagem
    imagem_label.pack(pady=20)

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

app = ctk.CTk()
app.title("Controle de Débitos Municipais")

# Definindo a geometria da janela
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
app.geometry(f"{screen_width}x{screen_height-40}+0+0")

# Título da janela
title_label = ctk.CTkLabel(
    app,
    text="Controle de Débitos Municipais",
    font=("Helvetica", 30, "bold"),
    text_color="#ffffff"
)
title_label.pack(pady=20)

# Frame principal
main_frame = ctk.CTkFrame(app, fg_color="transparent")
main_frame.pack(padx=20, pady=20, fill="both", expand=True)

# Adicionando campos de entrada e botão de iniciar
entry_sao_paulo = ctk.CTkEntry(
    main_frame, 
    placeholder_text="Digite a linha inicial do Excel para São Paulo", 
    font=("Helvetica", 14), 
    width=300
)
entry_sao_paulo.pack(pady=10)

exibir_imagem()

ultima_linha_processada_sao_paulo = ler_progresso_sao_paulo()

ultima_sao_paulo = ctk.CTkLabel(
    app,
    text=f"(Última linha processada: {ultima_linha_processada_sao_paulo})",
    font=("Helvetica", 12),
    text_color="lightgray"
)
ultima_sao_paulo.pack(pady=10)

atualizar_ultima_linha()

# Resultado do processamento
resultado_label = ctk.CTkLabel(
    app,
    text="Aguardando início...",
    font=("Helvetica", 14),
    text_color="yellow"
)
resultado_label.pack(pady=10)

# Botão para iniciar o processamento
iniciar_button = ctk.CTkButton(
    app,
    text="Iniciar Processamento",
    command=iniciar_processo,
    font=("Helvetica", 14),
    width=300,
    height=40,
    fg_color="#007acc",
    hover_color="#005b99"
)
iniciar_button.pack(pady=20)


# Rodapé
footer_label = ctk.CTkLabel(
    app,
    text="Desenvolvido por Luiz Fernando Hillebrande",
    font=("Helvetica", 10),
    text_color="#c9c9c9"
)
footer_label.pack(side="bottom", pady=25)

# Função para sair da tela cheia
def sair_tela_cheia(event=None):
    app.attributes('-fullscreen', False)

fechar_button = ctk.CTkButton(app, text="Fechar", command=fechar_janela)
fechar_button.pack(pady=10)

# Configura o fechamento da janela usando o protocolo WM_DELETE_WINDOW
app.protocol("WM_DELETE_WINDOW", fechar_janela)


app.bind("<Escape>", sair_tela_cheia)

# Inicia o loop da interface gráfica
app.mainloop()